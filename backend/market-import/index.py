"""
Универсальный импортёр рыночных данных в market_listings.
Поддерживает CSV (выгрузка парсера: ЦИАН, Авито, Яндекс) и XLSX (ручная выгрузка).

Формат CSV: разделитель ";", кодировка UTF-8/UTF-8-BOM
  Колонки: Название, Цена, Дата, Телефон, Метро/Район, Адрес, Тип объявления,
           Источник, lat, lng, Доп.параметры (Этаж=N|Этажность=N|Вид объекта=X|Общая площадь=N)

Формат XLSX: произвольные колонки, обязательные: Цена, Площадь, Тип сделки.

Архитектура (батчевая, без 504):
  import_start    → создаёт job, возвращает job_id (~100 мс)
  import_continue → скачивает файл потоково, пропускает checkpoint строк,
                    парсит и вставляет BATCH строк, возвращает новый checkpoint
                    Фронтенд вызывает в цикле до done=true.
"""

import csv
import io
import json
import os
import re
import urllib.request
from datetime import datetime

import psycopg2
import psycopg2.extras

SCHEMA = os.environ.get('MAIN_DB_SCHEMA', 't_p71821556_real_estate_catalog_')
BATCH = 1500  # строк за один вызов import_continue

CORS = {
    'Access-Control-Allow-Origin': '*',
    'Access-Control-Allow-Methods': 'GET, POST, OPTIONS',
    'Access-Control-Allow-Headers': 'Content-Type, X-Auth-Token, X-Authorization',
}

# ── Маппинги ──────────────────────────────────────────────────────────────────

OBJ_TYPE_MAP = {
    'офисное помещение': 'office', 'офис': 'office',
    'торговое помещение': 'retail', 'торговый': 'retail', 'магазин': 'retail',
    'помещение свободного назначения': 'free_purpose', 'свободного назначения': 'free_purpose',
    'складское помещение': 'warehouse', 'склад': 'warehouse',
    'производственное помещение': 'production', 'производство': 'production',
    'здание': 'building', 'отдельно стоящее здание': 'building',
    'помещение общепита': 'restaurant', 'общепит': 'restaurant',
    'кафе': 'restaurant', 'ресторан': 'restaurant',
    'гостиница': 'hotel', 'апартаменты': 'hotel',
    'коммерческая земля': 'land', 'земля': 'land',
    'автосервис': 'car_service', 'автомойка': 'car_service',
    'готовый арендный бизнес': 'gab',
    'другое': 'other', 'иное': 'other',
}

DEAL_MAP = {
    'продам': 'sale', 'продажа': 'sale', 'продаётся': 'sale', 'продается': 'sale',
    'сдам': 'rent', 'аренда': 'rent', 'сдаётся': 'rent', 'сдается': 'rent',
}

DISTRICT_NORM = {
    'р-н прикубанский': 'Прикубанский', 'прикубанский': 'Прикубанский',
    'р-н карасунский': 'Карасунский', 'карасунский': 'Карасунский',
    'р-н западный': 'Западный', 'западный': 'Западный',
    'р-н центральный': 'Центральный', 'центральный': 'Центральный',
    'р-н прикубанский округ': 'Прикубанский',
}

MIN_PRICE_SALE = 100_000
MAX_PRICE_SALE = 5_000_000_000
MIN_PRICE_RENT = 3_000
MAX_PRICE_RENT = 10_000_000
MIN_AREA = 4
MAX_AREA = 100_000
FRESH_DAYS = 365

GAB_SIGNALS = ['с арендатором', 'арендный бизнес', 'готовый арендный', 'арендный поток']


# ── Утилиты ───────────────────────────────────────────────────────────────────

def _get_conn():
    return psycopg2.connect(os.environ['DATABASE_URL'], cursor_factory=psycopg2.extras.RealDictCursor)


def _map_obj_type(raw: str, title: str = '') -> str:
    s = (raw or '').lower().strip()
    cat = next((v for k, v in OBJ_TYPE_MAP.items() if k in s), 'other')
    if any(sig in (title or '').lower() for sig in GAB_SIGNALS):
        return 'gab'
    return cat


def _map_deal(raw: str) -> str:
    s = (raw or '').lower().strip()
    return next((v for k, v in DEAL_MAP.items() if k in s), 'sale')


def _norm_district(raw: str) -> str:
    if not raw:
        return ''
    return DISTRICT_NORM.get(raw.lower().strip(), raw.strip())


def _parse_float(s) -> float:
    try:
        return float(str(s).replace(' ', '').replace(',', '.').replace('\xa0', ''))
    except Exception:
        return 0.0


def _valid_date(date_str: str) -> bool:
    try:
        return (datetime.now() - datetime.fromisoformat(date_str[:10])).days <= FRESH_DAYS
    except Exception:
        return True


def _job_upsert(conn, job_id, **kwargs):
    sets = ', '.join(f"{k}=%s" for k in kwargs)
    vals = list(kwargs.values()) + [job_id]
    cur = conn.cursor()
    cur.execute(f"UPDATE {SCHEMA}.import_jobs SET {sets}, updated_at=NOW() WHERE id=%s", vals)
    conn.commit()
    cur.close()


# ── Парсинг одной строки CSV → запись ─────────────────────────────────────────

def _parse_csv_row(row: dict, source: str, row_num: int) -> tuple[dict | None, str | None]:
    """Парсит одну строку CSV. Возвращает (запись, предупреждение) или (None, причина_пропуска)."""
    title     = (row.get('Название') or '').strip()
    raw_price = (row.get('Цена') or '').strip().replace(' ', '').replace('\xa0', '')
    date_str  = (row.get('Дата') or '')[:19]
    phone     = (row.get('Телефон') or '').strip()
    district  = _norm_district(row.get('Метро/Район') or '')
    address   = (row.get('Адрес') or '').strip()
    deal_raw  = (row.get('Тип объявления') or '')
    src       = (row.get('Источник') or source or 'csv').strip()
    ext_id    = str(row.get('ID на сайте') or '').strip()
    url       = (row.get('URL') or '').strip()
    extra     = (row.get('Доп.параметры') or '')

    if date_str and not _valid_date(date_str):
        return None, f'row {row_num}: устаревшее'

    try:
        price = float(raw_price) if raw_price else 0.0
    except Exception:
        price = 0.0

    deal = _map_deal(deal_raw)

    if deal == 'sale' and not (MIN_PRICE_SALE <= price <= MAX_PRICE_SALE):
        return None, None
    if deal == 'rent' and not (MIN_PRICE_RENT <= price <= MAX_PRICE_RENT):
        return None, None

    area = 0.0
    m = re.search(r'Общая площадь=([0-9.,]+)', extra)
    if m:
        area = _parse_float(m.group(1))
    if area <= 0:
        m2 = re.search(r'\((\d+[\.,]?\d*)\s*м', title)
        if m2:
            area = _parse_float(m2.group(1))

    if not (MIN_AREA <= area <= MAX_AREA):
        return None, None

    ot_m = re.search(r'Вид объекта=([^|]+)', extra)
    obj_type_raw = ot_m.group(1).strip() if ot_m else ''
    category = _map_obj_type(obj_type_raw, title=title)

    floor, total_floors = None, None
    fl_m = re.search(r'Этаж=(\d+)', extra)
    tf_m = re.search(r'Этажность здания=(\d+)', extra)
    if fl_m:
        floor = int(fl_m.group(1))
    if tf_m:
        total_floors = int(tf_m.group(1))

    lat_s = (row.get('lat') or '').strip()
    lng_s = (row.get('lng') or '').strip()
    lat = _parse_float(lat_s)
    lng = _parse_float(lng_s)
    if not (44.0 < lat < 45.7 and 38.5 < lng < 39.5):
        lat, lng = 0.0, 0.0

    road_line = None
    rl_m = re.search(r'Линия улицы=([^|]+)', extra)
    if rl_m:
        road_line = rl_m.group(1).strip()[:50]

    return {
        'source': src[:50],
        'external_id': ext_id or None,
        'url': url or None,
        'title': title[:500] or None,
        'category': category,
        'deal_type': deal,
        'price': int(price) if price else None,
        'price_per_m2': round(price / area, 2) if price and area > 0 else None,
        'area': area or None,
        'address': address[:500] or None,
        'district': district[:200] or None,
        'floor': floor,
        'total_floors': total_floors,
        'phone': phone[:50] or None,
        'lat': lat or None,
        'lng': lng or None,
        'road_line': road_line,
    }, None


# ── Парсинг одной строки XLSX → запись ────────────────────────────────────────

def _parse_xlsx_row(row: tuple, header_idx: dict, source: str) -> dict | None:
    """Парсит одну строку XLSX. Возвращает запись или None."""
    def cell(name):
        idx = header_idx.get(name, -1)
        if idx < 0 or idx >= len(row):
            return None
        return row[idx]

    price = _parse_float(cell('price'))
    area  = _parse_float(cell('area'))
    if area <= 0:
        title_v = str(cell('title') or '')
        m = re.search(r'(\d+[\.,]?\d*)\s*м', title_v)
        if m:
            area = _parse_float(m.group(1))

    deal_raw = str(cell('deal') or 'продажа')
    deal = _map_deal(deal_raw)

    if deal == 'sale' and not (MIN_PRICE_SALE <= price <= MAX_PRICE_SALE):
        return None
    if deal == 'rent' and not (MIN_PRICE_RENT <= price <= MAX_PRICE_RENT):
        return None
    if not (MIN_AREA <= area <= MAX_AREA):
        return None

    cat_raw  = str(cell('cat') or '')
    title_v  = str(cell('title') or '').strip()
    category = _map_obj_type(cat_raw, title=title_v)
    address  = str(cell('addr') or '').strip()
    district = _norm_district(str(cell('dist') or ''))

    floor_v   = int(_parse_float(cell('floor'))) or None
    tfloors_v = int(_parse_float(cell('tfloors'))) or None
    url_v     = str(cell('url') or '').strip() or None
    ext_id    = str(cell('ext_id') or '').strip() or None
    ppm2_v    = _parse_float(cell('ppm2'))
    if not ppm2_v and area > 0 and price > 0:
        ppm2_v = round(price / area, 2)

    return {
        'source': source[:50],
        'external_id': ext_id,
        'url': url_v,
        'title': title_v[:500] or None,
        'category': category,
        'deal_type': deal,
        'price': int(price) if price else None,
        'price_per_m2': ppm2_v or None,
        'area': area or None,
        'address': address[:500] or None,
        'district': district[:200] or None,
        'floor': floor_v,
        'total_floors': tfloors_v,
        'phone': None,
        'lat': None,
        'lng': None,
        'road_line': None,
    }


def _build_xlsx_header_idx(header_row: tuple) -> dict:
    """Сопоставляет колонки XLSX → внутренние имена."""
    header = [str(c).strip().lower() if c else '' for c in header_row]

    def find(*candidates):
        for cand in candidates:
            for i, h in enumerate(header):
                if cand in h:
                    return i
        return -1

    return {
        'price':   find('цена', 'price', 'стоимость'),
        'area':    find('площадь', 'area', 'кв.м', 'кв м', 'площ', 'square'),
        'deal':    find('тип сделки', 'тип объявления', 'deal', 'сделка', 'операция'),
        'cat':     find('категория1', 'категория', 'тип объекта', 'вид объекта', 'category', 'назначение'),
        'addr':    find('адрес', 'address', 'местоположение'),
        'dist':    find('метро/район', 'район', 'district', 'округ'),
        'title':   find('название', 'заголовок', 'title', 'наименование'),
        'floor':   find('этаж', 'floor'),
        'tfloors': find('этажность', 'этажей', 'total_floor'),
        'url':     find('url', 'ссылка', 'link'),
        'ext_id':  find('id на сайте', 'id объявления', 'внешний id'),
        'ppm2':    find('цена за м', 'price_per_m', 'цена/м', 'руб/м'),
    }


# ── Вставка батча в БД ────────────────────────────────────────────────────────

def _insert_batch(cur, conn, records: list[dict]) -> tuple[int, int, int]:
    """Вставляет записи батчами, возвращает (inserted, updated, skipped)."""
    if not records:
        return 0, 0, 0

    inserted = updated = skipped = 0
    esc = lambda v: str(v).replace("'", "''") if v else ''

    # Дедупликация внутри батча по external_id
    sources = list({r['source'] for r in records if r.get('source')})
    extids  = [r['external_id'] for r in records if r.get('external_id')]

    existing = set()
    if extids and sources:
        src_list = ','.join(f"'{esc(s)}'" for s in sources)
        eid_list = ','.join(f"'{esc(e)}'" for e in extids[:5000])
        cur.execute(f"""
            SELECT external_id FROM {SCHEMA}.market_listings
            WHERE source IN ({src_list}) AND external_id IN ({eid_list})
        """)
        existing = {row['external_id'] for row in cur.fetchall()}

    to_insert, to_update = [], []
    for r in records:
        eid = r.get('external_id') or ''
        if eid and eid in existing:
            to_update.append(r)
        else:
            to_insert.append(r)

    # Батчевый INSERT по 200 строк
    for i in range(0, len(to_insert), 200):
        batch = to_insert[i:i+200]
        vals = []
        for r in batch:
            def s(v): return f"'{esc(str(v))}'" if v is not None else 'NULL'
            def n(v): return str(v) if v is not None else 'NULL'
            vals.append(
                f"({s(r['source'])},{s(r.get('external_id'))},{s(r.get('url'))},"
                f"{s(r.get('title'))},{s(r['category'])},{s(r['deal_type'])},"
                f"{n(r.get('price'))},{n(r.get('price_per_m2'))},{n(r.get('area'))},"
                f"{s(r.get('address'))},{s(r.get('district'))},{n(r.get('floor'))},"
                f"{n(r.get('total_floors'))},{s(r.get('phone'))},NOW())"
            )
        if vals:
            cur.execute(
                f"INSERT INTO {SCHEMA}.market_listings "
                f"(source,external_id,url,title,category,deal_type,price,price_per_m2,"
                f"area,address,district,floor,total_floors,phone,scraped_at) "
                f"VALUES {','.join(vals)} ON CONFLICT DO NOTHING"
            )
            inserted += cur.rowcount

    # Батчевый UPDATE
    for r in to_update:
        eid = esc(r.get('external_id') or '')
        src = esc(r.get('source') or '')
        ps  = str(r['price'])       if r.get('price')       is not None else 'NULL'
        pp  = str(r['price_per_m2'])if r.get('price_per_m2')is not None else 'NULL'
        ar  = str(r['area'])        if r.get('area')         is not None else 'NULL'
        cur.execute(
            f"UPDATE {SCHEMA}.market_listings SET price={ps},price_per_m2={pp},"
            f"area={ar},scraped_at=NOW() "
            f"WHERE source='{src}' AND external_id='{eid}'"
        )
        if cur.rowcount:
            updated += 1
        else:
            skipped += 1

    conn.commit()
    return inserted, updated, skipped


# ── Фоновые триггеры ──────────────────────────────────────────────────────────

def _trigger_aggregate():
    import threading, sys as _sys
    def _run():
        try:
            _pp = os.path.join(os.path.dirname(__file__), '..', 'price-predict')
            if _pp not in _sys.path:
                _sys.path.insert(0, _pp)
            from market_snapshots import aggregate_market_listings as _aml
            import psycopg2 as _pg2, psycopg2.extras as _ext
            _c = _pg2.connect(os.environ['DATABASE_URL'], cursor_factory=_ext.RealDictCursor)
            try:
                r = _aml(_c.cursor(), _c)
                print(f'[market-import] bg aggregate: {r}')
            finally:
                _c.close()
        except Exception as e:
            print(f'[market-import] bg aggregate error: {e}')
    threading.Thread(target=_run, daemon=True).start()


def _trigger_vb_retrain(cur):
    import threading
    cur.execute(f"""
        SELECT s.token FROM {SCHEMA}.sessions s
        JOIN {SCHEMA}.users u ON u.id = s.user_id
        WHERE u.role='admin' AND s.expires_at > NOW()
        ORDER BY s.created_at DESC LIMIT 1
    """)
    row = cur.fetchone()
    if not row:
        return False, 'no admin session'
    token = row['token']
    retrain_url = os.environ.get('RETRAIN_URL', '')
    if not retrain_url:
        return False, 'RETRAIN_URL not set'
    def _fire():
        try:
            payload = json.dumps({'sources': ['market_import']}).encode()
            req = urllib.request.Request(retrain_url, data=payload,
                headers={'Content-Type': 'application/json', 'X-Auth-Token': token},
                method='POST')
            with urllib.request.urlopen(req, timeout=60) as r:
                r.read()
        except Exception as e:
            print(f'[market-import] retrain error: {e}')
    threading.Thread(target=_fire, daemon=True).start()
    return True, None


# ── Handler ───────────────────────────────────────────────────────────────────

def handler(event: dict, context) -> dict:
    """Батчевый импорт рыночных данных из CSV или XLSX в market_listings."""
    if event.get('httpMethod') == 'OPTIONS':
        return {'statusCode': 200, 'headers': CORS, 'body': ''}

    def ok(data):
        return {'statusCode': 200, 'headers': {**CORS, 'Content-Type': 'application/json'},
                'body': json.dumps(data, ensure_ascii=False, default=str)}

    def err(code, msg):
        return {'statusCode': code, 'headers': {**CORS, 'Content-Type': 'application/json'},
                'body': json.dumps({'error': msg}, ensure_ascii=False)}

    body = {}
    if event.get('body'):
        try:
            body = json.loads(event['body'])
        except Exception:
            pass
    qs     = event.get('queryStringParameters') or {}
    action = body.get('action') or qs.get('action') or 'import'

    conn = _get_conn()
    cur  = conn.cursor()

    try:
        # ── Статистика ────────────────────────────────────────────────────────
        if action == 'stats':
            cur.execute(f"""
                SELECT source, deal_type, category,
                  COUNT(*) as cnt,
                  ROUND(AVG(price_per_m2)::numeric, 0) as avg_ppm2,
                  MAX(scraped_at) as last_scraped
                FROM {SCHEMA}.market_listings
                GROUP BY source, deal_type, category
                ORDER BY source, deal_type, cnt DESC
            """)
            rows = cur.fetchall()
            cur.execute(f"SELECT COUNT(*) as total FROM {SCHEMA}.market_listings")
            total = cur.fetchone()['total']
            return ok({'total': total, 'breakdown': [dict(r) for r in rows]})

        # ── Очистка по источнику ──────────────────────────────────────────────
        if action == 'clear':
            source = (body.get('source') or '').strip()
            if not source:
                return err(400, 'Укажите source')
            src_safe = source.replace("'", "''")
            cur.execute(f"DELETE FROM {SCHEMA}.market_listings WHERE source='{src_safe}'")
            deleted = max(0, cur.rowcount)
            conn.commit()
            return ok({'deleted': deleted, 'source': source})

        # ── Предпросмотр — скачать+парсить первые 200 строк (~5 сек) ─────────
        if action in ('import_preview', 'import') and body.get('preview'):
            file_url = body.get('file_url') or qs.get('file_url') or ''
            source   = body.get('source') or 'manual'
            if not file_url:
                return err(400, 'Укажите file_url')
            req = urllib.request.Request(file_url, headers={'User-Agent': 'Mozilla/5.0'})
            with urllib.request.urlopen(req, timeout=25) as resp:
                raw = resp.read(5_000_000)  # читаем только первые 5 МБ для превью
            file_lower = file_url.lower().split('?')[0]
            is_xlsx = file_lower.endswith('.xlsx') or file_lower.endswith('.xls')
            records = []
            if is_xlsx:
                import openpyxl
                wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
                ws = wb.active
                rows_iter = ws.iter_rows(values_only=True)
                header_row = next(rows_iter, None)
                if header_row:
                    hidx = _build_xlsx_header_idx(header_row)
                    for row in rows_iter:
                        rec = _parse_xlsx_row(row, hidx, source)
                        if rec:
                            records.append(rec)
                        if len(records) >= 200:
                            break
                wb.close()
            else:
                text = raw.decode('utf-8-sig', errors='replace')
                reader = csv.DictReader(io.StringIO(text), delimiter=';')
                for i, row in enumerate(reader, 1):
                    rec, _ = _parse_csv_row(row, source, i)
                    if rec:
                        records.append(rec)
                    if len(records) >= 200:
                        break
            by_cat = {}; by_deal = {}
            for r in records:
                by_cat[r['category']] = by_cat.get(r['category'], 0) + 1
                by_deal[r['deal_type']] = by_deal.get(r['deal_type'], 0) + 1
            prices = [r['price'] for r in records if r.get('price')]
            areas  = [float(r['area']) for r in records if r.get('area')]
            import statistics as _st
            return ok({
                'preview': True,
                'format': 'xlsx' if is_xlsx else 'csv',
                'total_parsed': len(records),
                'warnings_count': 0,
                'warnings_sample': [],
                'by_category': by_cat,
                'by_deal': by_deal,
                'price_median': round(_st.median(prices)) if prices else None,
                'area_median': round(_st.median(areas)) if areas else None,
                'sample': records[:10],
            })

        # ── Старт импорта — только создаём job (~50 мс) ───────────────────────
        if action in ('import_start', 'import', 'import_market_start'):
            file_url = body.get('file_url') or qs.get('file_url') or ''
            source   = body.get('source') or 'manual'
            replace  = bool(body.get('replace', False))
            if not file_url:
                return err(400, 'Укажите file_url')

            # Очищаем старые данные если нужно
            if replace:
                src_safe = source.replace("'", "''")
                cur.execute(f"DELETE FROM {SCHEMA}.market_listings WHERE source='{src_safe}'")
                conn.commit()

            cur.execute(
                f"INSERT INTO {SCHEMA}.import_jobs "
                f"(file_url, source, replace_existing, status, checkpoint_row) "
                f"VALUES (%s, %s, %s, 'running', 0) RETURNING id",
                (file_url, source, replace)
            )
            job_id = cur.fetchone()['id']
            conn.commit()
            return ok({'job_id': job_id, 'status': 'running', 'done': False})

        # ── Продолжение — скачать, пропустить до checkpoint, вставить BATCH строк
        if action in ('import_continue', 'import_market_continue'):
            job_id = body.get('job_id')
            if not job_id:
                return err(400, 'job_id required')

            cur.execute(f"SELECT * FROM {SCHEMA}.import_jobs WHERE id=%s", (job_id,))
            job = cur.fetchone()
            if not job:
                return err(404, 'job not found')
            if job['status'] != 'running':
                return ok(dict(job))

            file_url   = job['file_url']
            source     = job['source']
            checkpoint = job['checkpoint_row'] or 0
            rows_ins   = job['rows_inserted'] or 0
            rows_upd   = job['rows_updated'] or 0
            rows_skip  = job['rows_skipped'] or 0
            cat_counts = dict(job.get('category_breakdown') or {})

            file_lower = file_url.lower().split('?')[0]
            is_xlsx = file_lower.endswith('.xlsx') or file_lower.endswith('.xls')

            # Скачиваем файл (каждый раз — CDN отдаёт быстро из кеша)
            req = urllib.request.Request(file_url, headers={'User-Agent': 'Mozilla/5.0'})
            with urllib.request.urlopen(req, timeout=25) as resp:
                raw = resp.read()

            # Собираем записи: пропускаем до checkpoint, берём следующие BATCH
            batch_records = []
            total_rows = 0
            warnings = []
            reached_end = False

            if is_xlsx:
                import openpyxl
                wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
                ws = wb.active
                rows_iter = ws.iter_rows(values_only=True)
                header_row = next(rows_iter, None)
                if not header_row:
                    wb.close()
                    _job_upsert(conn, job_id, status='error', error_msg='Файл пустой')
                    return err(400, 'Файл пустой')
                hidx = _build_xlsx_header_idx(header_row)
                seen = set()
                for row in rows_iter:
                    total_rows += 1
                    if total_rows <= checkpoint:
                        continue
                    rec = _parse_xlsx_row(row, hidx, source)
                    if rec:
                        dk = f"{(rec.get('address') or '')}|{int(rec.get('area') or 0)}"
                        if dk not in seen:
                            seen.add(dk)
                            batch_records.append(rec)
                    if len(batch_records) >= BATCH:
                        break
                else:
                    reached_end = True
                if not reached_end and total_rows < checkpoint + BATCH:
                    reached_end = True
                wb.close()
            else:
                text = raw.decode('utf-8-sig', errors='replace')
                reader = csv.DictReader(io.StringIO(text), delimiter=';')
                seen = set()
                for i, row in enumerate(reader, 1):
                    total_rows = i
                    if i <= checkpoint:
                        continue
                    rec, warn = _parse_csv_row(row, source, i)
                    if warn:
                        warnings.append(warn)
                    if rec:
                        dk = f"{(rec.get('address') or '')}|{int(rec.get('area') or 0)}"
                        if dk not in seen:
                            seen.add(dk)
                            batch_records.append(rec)
                    if len(batch_records) >= BATCH:
                        break
                else:
                    reached_end = True

            # Вставляем батч
            ins, upd, skp = _insert_batch(cur, conn, batch_records)
            new_checkpoint = checkpoint + (total_rows - checkpoint)
            new_ins  = rows_ins + ins
            new_upd  = rows_upd + upd
            new_skip = rows_skip + skp

            for r in batch_records:
                cat = r.get('category', 'other')
                cat_counts[cat] = cat_counts.get(cat, 0) + 1

            done = reached_end and len(batch_records) < BATCH

            _job_upsert(conn, job_id,
                status='done' if done else 'running',
                rows_done=new_checkpoint,
                rows_inserted=new_ins,
                rows_updated=new_upd,
                rows_skipped=new_skip,
                checkpoint_row=new_checkpoint,
                category_breakdown=json.dumps(cat_counts, ensure_ascii=False))

            if done and new_ins > 0:
                _trigger_vb_retrain(cur)
                _trigger_aggregate()

            return ok({
                'done': done,
                'job_id': job_id,
                'rows_done': new_checkpoint,
                'rows_total': total_rows,
                'rows_inserted': new_ins,
                'rows_updated': new_upd,
                'inserted': new_ins,
                'updated': new_upd,
                'status': 'done' if done else 'running',
                'category_breakdown': cat_counts,
            })

        # ── Алиасы xlsx-reader (import_market_status/list) ───────────────────────
        if action in ('import_market_status',):
            job_id = body.get('job_id')
            if not job_id:
                return err(400, 'job_id required')
            cur.execute(f"SELECT * FROM {SCHEMA}.import_jobs WHERE id=%s", (job_id,))
            row = cur.fetchone()
            if not row:
                return err(404, 'job not found')
            return ok(dict(row))

        if action in ('import_market_list',):
            cur.execute(f"SELECT * FROM {SCHEMA}.import_jobs ORDER BY created_at DESC LIMIT 20")
            return ok([dict(r) for r in cur.fetchall()])

        return err(400, f'Неизвестный action: {action}')

    finally:
        cur.close()
        conn.close()