"""
adapters/xlsx.py — адаптер импорта XLSX-файлов.

Поддерживает два режима:
1. Потоковый (job-система) — для больших файлов (до 200 МБ), запускается фоном
   через threading. Прогресс пишется в import_jobs.
2. Прямой (синхронный) — preview + небольшие файлы, до 10 000 строк.

Колонки обнаруживаются автоматически по нечёткому совпадению заголовков.
Все маппинги и утилиты — из core.py.
"""

import io
import json
import os
import re
import threading

import openpyxl
import psycopg2
import psycopg2.extras

from core import (
    SCHEMA,
    parse_numeric, parse_area, ppm2,
    map_category, map_deal, norm_district,
    validate_record, valid_date, valid_coords, dedup_key,
    fetch_bytes, upsert_batch, job_update,
)

BATCH_SIZE = 500  # строк за один flush в БД


# ═══════════════════════════════════════════════════════════════════════════════
# ОПРЕДЕЛЕНИЕ КОЛОНОК
# ═══════════════════════════════════════════════════════════════════════════════

def _find_col(header: list[str], *candidates: str) -> int:
    """Нечёткий поиск колонки по заголовку. Возвращает индекс или -1."""
    for cand in candidates:
        for i, h in enumerate(header):
            if cand in h:
                return i
    return -1


def _detect_cols(header: list[str]) -> dict:
    """
    Определяет индексы всех нужных колонок из заголовка.
    Возвращает словарь {имя: индекс}.
    """
    f = lambda *c: _find_col(header, *c)
    return {
        'price':    f('цена', 'price', 'стоимость'),
        'area':     f('площадь', 'area', 'кв.м', 'кв м', 'площ', 'square'),
        'deal':     f('тип сделки', 'тип объявления', 'deal', 'сделка', 'операция'),
        'cat':      f('категория1', 'категория', 'тип объекта', 'вид объекта', 'category',
                      'тип недвижимости', 'назначение'),
        'cat2':     f('категория2'),
        'addr':     f('адрес', 'address', 'местоположение'),
        'dist':     f('метро/район', 'район', 'district', 'округ', 'метро'),
        'title':    f('название', 'заголовок', 'title', 'наименование'),
        'floor':    f('этаж', 'floor'),
        'tfloors':  f('этажность', 'этажей', 'total_floor', 'кол-во этажей'),
        'url':      f('url', 'ссылка', 'link', 'объявление'),
        'source':   f('источник'),
        'ext_id':   f('id на сайте', 'id объявления', 'внешний id', 'id', 'номер объявления'),
        'desc':     f('описание', 'description', 'комментарий'),
        'ppm2':     f('цена за м', 'price_per_m', 'цена/м', 'руб/м'),
        'phone':    f('телефон', 'phone', 'контакт'),
        'lat':      f('lat', 'latitude', 'широта'),
        'lng':      f('lng', 'lon', 'longitude', 'долгота'),
        'date':     f('дата', 'date', 'опубликован'),
    }


# ═══════════════════════════════════════════════════════════════════════════════
# ПАРСИНГ ОДНОЙ СТРОКИ
# ═══════════════════════════════════════════════════════════════════════════════

def _cell(row, idx):
    """Безопасное получение ячейки по индексу."""
    if idx < 0 or idx >= len(row):
        return None
    return row[idx]


def _parse_row(row, cols: dict, source: str, row_num: int) -> tuple[dict | None, str]:
    """
    Парсит одну строку XLSX → словарь записи.
    Возвращает (запись, '') или (None, причина_пропуска).
    """
    c = lambda name: _cell(row, cols[name])

    title_v = str(c('title') or '').strip()
    desc_v  = str(c('desc')  or '').strip() if cols.get('desc', -1) >= 0 else ''

    # Цена
    price = int(parse_numeric(c('price')))

    # Площадь: из колонки → из названия → из описания
    area = parse_area(c('area'), title=title_v, description=desc_v)

    # Тип сделки
    deal = map_deal(str(c('deal') or 'продажа'))

    # Категория: объединяем cat + cat2 (Авито присылает двухуровневую)
    cat_raw = str(c('cat') or '').strip()
    if cols.get('cat2', -1) >= 0:
        cat2 = str(_cell(row, cols['cat2']) or '').strip()
        cat_raw = f'{cat_raw} {cat2}'.strip()
    category = map_category(cat_raw, title=title_v)

    # Прочие поля
    address   = str(c('addr')    or '').strip()
    district  = norm_district(str(c('dist') or '').strip(), address)
    floor_v   = int(parse_numeric(c('floor')))    or None
    tfloors_v = int(parse_numeric(c('tfloors')))  or None
    url_v     = str(c('url')     or '').strip() or None
    phone_v   = str(c('phone')   or '').strip() or None
    desc_out  = desc_v[:1000] or None

    # Источник: из колонки или переданный параметром
    row_source = str(_cell(row, cols.get('source', -1)) or source or 'xlsx')[:50]

    # external_id
    ext_id = str(c('ext_id') or '').strip() or None

    # Цена за м² — из колонки или вычисляем
    ppm2_raw = parse_numeric(c('ppm2')) if cols.get('ppm2', -1) >= 0 else 0.0
    ppm2_val = ppm2_raw if ppm2_raw else ppm2(price, area)

    # Координаты
    lat_v = parse_numeric(c('lat')) if cols.get('lat', -1) >= 0 else None
    lng_v = parse_numeric(c('lng')) if cols.get('lng', -1) >= 0 else None
    if lat_v and lng_v and not valid_coords(lat_v, lng_v):
        lat_v, lng_v = None, None

    # Дата публикации
    if cols.get('date', -1) >= 0:
        date_raw = str(c('date') or '')[:10]
        if date_raw and not valid_date(date_raw):
            return None, f'row {row_num}: устаревшее объявление ({date_raw})'

    rec = {
        'source':       row_source,
        'external_id':  ext_id,
        'url':          url_v,
        'title':        title_v[:500] or None,
        'category':     category,
        'deal_type':    deal,
        'price':        price or None,
        'price_per_m2': ppm2_val or None,
        'area':         area or None,
        'address':      address[:500] or None,
        'district':     district[:200] or None,
        'floor':        floor_v,
        'total_floors': tfloors_v,
        'phone':        phone_v[:50] if phone_v else None,
        'description':  desc_out,
        'lat':          lat_v or None,
        'lng':          lng_v or None,
    }

    ok, reason = validate_record(rec)
    if not ok:
        return None, f'row {row_num}: {reason}'

    return rec, ''


# ═══════════════════════════════════════════════════════════════════════════════
# СИНХРОННЫЙ РЕЖИМ (preview + небольшие файлы)
# ═══════════════════════════════════════════════════════════════════════════════

def parse_xlsx_sync(raw_bytes: bytes, source: str) -> tuple[list[dict], list[str], list[str]]:
    """
    Синхронный парсинг XLSX.
    Возвращает (записи, предупреждения, обнаруженные_колонки).
    Используется для preview и небольших файлов (< 10 000 строк).
    """
    wb = openpyxl.load_workbook(io.BytesIO(raw_bytes), read_only=True, data_only=True)
    ws = wb.active
    all_rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not all_rows:
        return [], ['Файл пустой'], []

    header = [str(c).strip().lower() if c else '' for c in all_rows[0]]
    cols = _detect_cols(header)

    if cols['price'] < 0:
        return [], [
            f'Не найдена обязательная колонка "Цена". '
            f'Найденные заголовки: {header[:20]}'
        ], header

    records, warnings = [], []
    seen = set()

    for i, row in enumerate(all_rows[1:], 2):
        rec, reason = _parse_row(row, cols, source, i)
        if rec is None:
            if reason:
                warnings.append(reason)
            continue
        dk = dedup_key(rec['source'], rec.get('external_id') or '',
                       rec.get('address') or f'row{i}',
                       rec.get('area') or 0, rec.get('price') or 0)
        if dk in seen:
            continue
        seen.add(dk)
        records.append(rec)

    detected = {k: header[v] for k, v in cols.items() if v >= 0}
    return records, warnings, list(detected.values())


# ═══════════════════════════════════════════════════════════════════════════════
# ФОНОВЫЙ ПОТОК (job-система для больших файлов)
# ═══════════════════════════════════════════════════════════════════════════════

def _run_job(job_id: int):
    """
    Фоновый поток: скачивает XLSX и импортирует батчами по BATCH_SIZE строк.
    Прогресс обновляется в import_jobs каждый батч.
    """
    conn = psycopg2.connect(os.environ['DATABASE_URL'])
    try:
        # Загружаем параметры задания
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute(f'SELECT * FROM {SCHEMA}.import_jobs WHERE id = %s', (job_id,))
        job = dict(cur.fetchone())
        cur.close()

        file_url = job['file_url']
        source   = job['source']
        replace  = job.get('replace_existing', False)

        # Шаг 1: скачиваем файл
        job_update(conn, job_id, status='downloading')
        raw_bytes = fetch_bytes(file_url, timeout=180)

        # Шаг 2: читаем заголовок
        job_update(conn, job_id, status='parsing')
        wb = openpyxl.load_workbook(io.BytesIO(raw_bytes), read_only=True, data_only=True)
        ws = wb.active

        rows_iter = ws.iter_rows(values_only=True)
        header_row = next(rows_iter, None)
        if not header_row:
            job_update(conn, job_id, status='error', error_msg='Файл пустой')
            wb.close()
            return

        header = [str(c).strip().lower() if c else '' for c in header_row]
        cols = _detect_cols(header)

        if cols['price'] < 0:
            job_update(conn, job_id, status='error',
                       error_msg=f'Не найдена колонка "Цена". Заголовки: {header[:20]}')
            wb.close()
            return

        # Если replace — удаляем старые записи этого источника
        if replace:
            cur2 = conn.cursor()
            cur2.execute(f"DELETE FROM {SCHEMA}.market_listings WHERE source = %s", (source,))
            conn.commit()
            cur2.close()

        # Шаг 3: импорт строк
        job_update(conn, job_id, status='running')

        rows_done = inserted = updated = skipped = 0
        cat_counts: dict[str, int] = {}
        seen: set[str] = set()
        batch: list[dict] = []

        def flush():
            nonlocal inserted, updated
            ins, upd = upsert_batch(conn, batch)
            inserted += ins
            updated  += upd

        for row in rows_iter:
            rows_done += 1

            rec, reason = _parse_row(row, cols, source, rows_done + 1)
            if rec is None:
                skipped += 1
                continue

            dk = dedup_key(rec['source'], rec.get('external_id') or '',
                           rec.get('address') or f'r{rows_done}',
                           rec.get('area') or 0, rec.get('price') or 0)
            if dk in seen:
                skipped += 1
                continue
            seen.add(dk)

            cat_counts[rec['category']] = cat_counts.get(rec['category'], 0) + 1
            batch.append(rec)

            if len(batch) >= BATCH_SIZE:
                flush()
                batch = []
                job_update(conn, job_id,
                           rows_done=rows_done,
                           rows_inserted=inserted,
                           rows_updated=updated,
                           rows_skipped=skipped)

        # Последний батч
        if batch:
            flush()

        wb.close()

        job_update(conn, job_id,
                   status='done',
                   rows_done=rows_done,
                   rows_inserted=inserted,
                   rows_updated=updated,
                   rows_skipped=skipped,
                   category_breakdown=json.dumps(cat_counts, ensure_ascii=False))

    except Exception as e:
        try:
            job_update(conn, job_id, status='error', error_msg=str(e)[:500])
        except Exception:
            pass
    finally:
        conn.close()


# ═══════════════════════════════════════════════════════════════════════════════
# ПУБЛИЧНЫЙ API АДАПТЕРА
# ═══════════════════════════════════════════════════════════════════════════════

def action_start(file_url: str, source: str, replace: bool) -> dict:
    """
    Создаёт job в import_jobs и запускает фоновый поток.
    Возвращает {job_id, status}.
    """
    conn = psycopg2.connect(os.environ['DATABASE_URL'])
    cur = conn.cursor()
    cur.execute(
        f"INSERT INTO {SCHEMA}.import_jobs (file_url, source, replace_existing, status) "
        f"VALUES (%s, %s, %s, 'pending') RETURNING id",
        (file_url, source, replace),
    )
    job_id = cur.fetchone()[0]
    conn.commit()
    cur.close()
    conn.close()

    t = threading.Thread(target=_run_job, args=(job_id,), daemon=True)
    t.start()

    return {'job_id': job_id, 'status': 'pending'}


def action_status(job_id: int) -> dict | None:
    """Возвращает текущий статус job или None если не найден."""
    conn = psycopg2.connect(
        os.environ['DATABASE_URL'],
        cursor_factory=psycopg2.extras.RealDictCursor,
    )
    cur = conn.cursor()
    cur.execute(f'SELECT * FROM {SCHEMA}.import_jobs WHERE id = %s', (job_id,))
    row = cur.fetchone()
    cur.close()
    conn.close()
    return dict(row) if row else None


def action_list(limit: int = 20) -> list[dict]:
    """Возвращает последние N заданий импорта."""
    conn = psycopg2.connect(
        os.environ['DATABASE_URL'],
        cursor_factory=psycopg2.extras.RealDictCursor,
    )
    cur = conn.cursor()
    cur.execute(
        f'SELECT * FROM {SCHEMA}.import_jobs ORDER BY created_at DESC LIMIT %s',
        (limit,),
    )
    rows = [dict(r) for r in cur.fetchall()]
    cur.close()
    conn.close()
    return rows


def action_preview(file_url: str, source: str) -> dict:
    """
    Скачивает файл и возвращает превью без записи в БД.
    Показывает обнаруженные колонки, статистику и 5 примеров.
    """
    raw_bytes = fetch_bytes(file_url, timeout=60)
    records, warnings, detected_cols = parse_xlsx_sync(raw_bytes, source)

    cat_counts: dict[str, int] = {}
    deal_counts: dict[str, int] = {}
    for r in records:
        cat_counts[r['category']]  = cat_counts.get(r['category'], 0) + 1
        deal_counts[r['deal_type']] = deal_counts.get(r['deal_type'], 0) + 1

    return {
        'preview':          True,
        'records_parsed':   len(records),
        'columns_detected': detected_cols,
        'category_breakdown': cat_counts,
        'deal_breakdown':   deal_counts,
        'sample':           records[:5],
        'warnings':         warnings[:20],
    }
