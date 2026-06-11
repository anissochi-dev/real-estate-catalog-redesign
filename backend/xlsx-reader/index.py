import json
import os
import io
import re
import urllib.request
import psycopg2
import psycopg2.extras
import openpyxl
from datetime import datetime

SCHEMA = os.environ.get('MAIN_DB_SCHEMA', 't_p71821556_real_estate_catalog_')

# ── Маппинги для import_market ────────────────────────────────────────────────

OBJ_TYPE_MAP = {
    'офисное помещение': 'office', 'офис': 'office',
    'торговое помещение': 'retail', 'торговый': 'retail', 'магазин': 'retail',
    'помещение свободного назначения': 'free_purpose', 'свободного назначения': 'free_purpose', 'псн': 'free_purpose',
    'складское помещение': 'warehouse', 'склад': 'warehouse',
    'производственное помещение': 'production', 'производство': 'production',
    'здание': 'building', 'отдельно стоящее здание': 'building',
    'помещение общепита': 'catering', 'общепит': 'catering', 'кафе': 'catering', 'ресторан': 'catering',
    'гостиница': 'hotel', 'апартаменты': 'hotel',
    'коммерческая земля': 'land', 'земля': 'land',
    'автосервис': 'car_service', 'автомойка': 'car_service',
    'готовый арендный бизнес': 'gab', 'габ': 'gab',
    'другое': 'other', 'иное': 'other',
}

DEAL_MAP = {
    'продам': 'sale', 'продажа': 'sale', 'продаётся': 'sale', 'продается': 'sale',
    'сдам': 'rent', 'аренда': 'rent', 'сдаётся': 'rent', 'сдается': 'rent',
}

GAB_TITLE_SIGNALS = ['с арендатором', 'арендный бизнес', 'готовый арендный', 'арендный поток']


def _map_obj_type_m(raw: str, title: str = '') -> str:
    s = (raw or '').lower().strip()
    cat = 'other'
    for key, val in OBJ_TYPE_MAP.items():
        if key in s:
            cat = val
            break
    t = (title or '').lower()
    if any(sig in t for sig in GAB_TITLE_SIGNALS):
        return 'gab'
    return cat


def _map_deal_m(raw: str) -> str:
    s = (raw or '').lower().strip()
    for key, val in DEAL_MAP.items():
        if key in s:
            return val
    return 'sale'


def _parse_float_m(v) -> float:
    if v is None:
        return 0.0
    s = re.sub(r'[^\d.,]', '', str(v)).replace(',', '.')
    try:
        return float(s)
    except Exception:
        return 0.0


def _ppm2_m(price: float, area: float) -> float | None:
    if price and area and area > 0:
        return round(price / area, 2)
    return None


def _import_market_xlsx(file_url: str, source: str, preview: bool, replace: bool) -> dict:
    """Скачивает большой XLSX (до 200 МБ) и импортирует в market_listings."""
    CORS = {'Access-Control-Allow-Origin': '*', 'Content-Type': 'application/json'}

    # Скачиваем файл (большой таймаут)
    try:
        req = urllib.request.Request(file_url, headers={'User-Agent': 'Mozilla/5.0'})
        with urllib.request.urlopen(req, timeout=120) as r:
            raw_bytes = r.read()
    except Exception as e:
        return {'statusCode': 500, 'headers': CORS, 'body': json.dumps({'error': f'Ошибка скачивания: {e}'}, ensure_ascii=False)}

    # Открываем XLSX
    try:
        wb = openpyxl.load_workbook(io.BytesIO(raw_bytes), read_only=True, data_only=True)
        ws = wb.active
        all_rows = list(ws.iter_rows(values_only=True))
        wb.close()
    except Exception as e:
        return {'statusCode': 500, 'headers': CORS, 'body': json.dumps({'error': f'Ошибка чтения XLSX: {e}'}, ensure_ascii=False)}

    if not all_rows:
        return {'statusCode': 400, 'headers': CORS, 'body': json.dumps({'error': 'Файл пустой'}, ensure_ascii=False)}

    header = [str(c).strip().lower() if c else '' for c in all_rows[0]]

    def find_col(*candidates) -> int:
        for cand in candidates:
            for i, h in enumerate(header):
                if cand in h:
                    return i
        return -1

    col_price   = find_col('цена', 'price', 'стоимость')
    col_area    = find_col('площадь', 'area', 'кв.м', 'кв м', 'площ')
    col_deal    = find_col('тип сделки', 'тип объявления', 'deal', 'сделка', 'операция')
    col_cat     = find_col('категория', 'тип объекта', 'вид объекта', 'category', 'тип недвижимости', 'назначение')
    col_addr    = find_col('адрес', 'address', 'местоположение')
    col_dist    = find_col('район', 'district', 'округ')
    col_title   = find_col('название', 'заголовок', 'title', 'наименование')
    col_floor   = find_col('этаж', 'floor')
    col_tfloors = find_col('этажность', 'этажей', 'total_floor', 'кол-во этажей')
    col_url     = find_col('url', 'ссылка', 'link', 'объявление')
    col_ext_id  = find_col('id объявления', 'внешний id', 'id', 'номер объявления')
    col_desc    = find_col('описание', 'description', 'комментарий')
    col_ppm2    = find_col('цена за м', 'price_per_m', 'цена/м', 'руб/м')

    if col_price < 0 or col_area < 0:
        return {
            'statusCode': 400, 'headers': CORS,
            'body': json.dumps({
                'error': 'Не найдены обязательные колонки "Цена" и "Площадь"',
                'header_found': header[:30],
            }, ensure_ascii=False)
        }

    records, warnings = [], []
    seen_keys = set()
    MIN_PRICE_SALE, MAX_PRICE_SALE = 500_000, 5_000_000_000
    MIN_PRICE_RENT, MAX_PRICE_RENT = 5_000, 10_000_000
    MIN_AREA, MAX_AREA = 1, 200_000

    for i, row in enumerate(all_rows[1:], 2):
        def cell(idx):
            if idx < 0 or idx >= len(row):
                return None
            return row[idx]

        price = _parse_float_m(cell(col_price))
        area  = _parse_float_m(cell(col_area))
        deal_raw = str(cell(col_deal) or 'продажа')
        deal = _map_deal_m(deal_raw)
        title_val = str(cell(col_title) or '').strip() if col_title >= 0 else ''
        cat_raw = str(cell(col_cat) or '').strip()
        category = _map_obj_type_m(cat_raw, title_val)
        address = str(cell(col_addr) or '').strip()
        district = str(cell(col_dist) or '').strip()[:200]
        floor_v = int(_parse_float_m(cell(col_floor))) or None
        tfloors_v = int(_parse_float_m(cell(col_tfloors))) or None
        url_v = str(cell(col_url) or '').strip() or None
        ext_id = str(cell(col_ext_id) or '').strip() or None
        desc = str(cell(col_desc) or '').strip()[:1000] or None
        ppm2_raw = _parse_float_m(cell(col_ppm2)) if col_ppm2 >= 0 else None
        ppm2 = ppm2_raw if ppm2_raw else _ppm2_m(price, area)

        if deal == 'sale' and not (MIN_PRICE_SALE <= price <= MAX_PRICE_SALE):
            if price > 0:
                warnings.append(f'row {i}: цена вне диапазона ({price:,.0f})')
            continue
        if deal == 'rent' and not (MIN_PRICE_RENT <= price <= MAX_PRICE_RENT):
            if price > 0:
                warnings.append(f'row {i}: цена аренды вне диапазона ({price:,.0f})')
            continue
        if not (MIN_AREA <= area <= MAX_AREA):
            if area > 0:
                warnings.append(f'row {i}: площадь вне диапазона ({area})')
            continue

        dk = f"{address or ext_id or f'row{i}'}_{int(area)}"
        if dk in seen_keys:
            continue
        seen_keys.add(dk)

        records.append({
            'source': (source or 'xlsx')[:50],
            'external_id': ext_id,
            'url': url_v,
            'title': title_val[:500] or None,
            'category': category,
            'deal_type': deal,
            'price': int(price) if price else None,
            'price_per_m2': ppm2,
            'area': area or None,
            'address': address[:500] or None,
            'district': district or None,
            'floor': floor_v,
            'total_floors': tfloors_v,
            'description': desc,
        })

    if preview:
        return {
            'statusCode': 200, 'headers': CORS,
            'body': json.dumps({
                'preview': True,
                'total_rows_in_file': len(all_rows) - 1,
                'records_parsed': len(records),
                'skipped': len(all_rows) - 1 - len(records),
                'columns_detected': {
                    'price': header[col_price] if col_price >= 0 else None,
                    'area': header[col_area] if col_area >= 0 else None,
                    'deal': header[col_deal] if col_deal >= 0 else None,
                    'category': header[col_cat] if col_cat >= 0 else None,
                    'address': header[col_addr] if col_addr >= 0 else None,
                    'district': header[col_dist] if col_dist >= 0 else None,
                    'url': header[col_url] if col_url >= 0 else None,
                    'ext_id': header[col_ext_id] if col_ext_id >= 0 else None,
                },
                'all_columns': header,
                'category_breakdown': _count_categories(records),
                'deal_breakdown': _count_deals(records),
                'sample': records[:5],
                'warnings': warnings[:20],
            }, ensure_ascii=False, default=str)
        }

    # Полный импорт в БД
    conn = psycopg2.connect(os.environ['DATABASE_URL'])
    cur = conn.cursor()
    inserted = updated = 0

    if replace:
        cur.execute(f"DELETE FROM {SCHEMA}.market_listings WHERE source = %s", (source,))
        conn.commit()

    for rec in records:
        ext_id = str(rec.get('external_id') or '')[:200] or f"xlsx_{rec.get('address','')[:50]}_{int(rec.get('area') or 0)}"
        cur.execute(
            f"INSERT INTO {SCHEMA}.market_listings "
            f"(source, external_id, url, title, category, deal_type, price, price_per_m2, "
            f"area, address, district, floor, total_floors, description, scraped_at) "
            f"VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,NOW()) "
            f"ON CONFLICT (source, external_id) DO UPDATE SET "
            f"price=%s, price_per_m2=%s, area=%s, category=%s, deal_type=%s, "
            f"address=%s, district=%s, title=%s, scraped_at=NOW()",
            (
                rec['source'], ext_id, rec.get('url'), rec.get('title'),
                rec['category'], rec['deal_type'], rec.get('price'), rec.get('price_per_m2'),
                rec.get('area'), rec.get('address'), rec.get('district'),
                rec.get('floor'), rec.get('total_floors'), rec.get('description'),
                rec.get('price'), rec.get('price_per_m2'), rec.get('area'), rec['category'], rec['deal_type'],
                rec.get('address'), rec.get('district'), rec.get('title'),
            )
        )
        if cur.rowcount == 1:
            inserted += 1
        else:
            updated += 1

    conn.commit()
    cur.close()
    conn.close()

    return {
        'statusCode': 200, 'headers': CORS,
        'body': json.dumps({
            'success': True,
            'total_rows_in_file': len(all_rows) - 1,
            'records_parsed': len(records),
            'inserted': inserted,
            'updated': updated,
            'warnings_count': len(warnings),
            'warnings': warnings[:10],
            'category_breakdown': _count_categories(records),
        }, ensure_ascii=False)
    }


def _count_categories(records):
    d = {}
    for r in records:
        d[r['category']] = d.get(r['category'], 0) + 1
    return d


def _count_deals(records):
    d = {}
    for r in records:
        d[r['deal_type']] = d.get(r['deal_type'], 0) + 1
    return d


def _job_update(conn, job_id: int, **kwargs):
    """Обновляет поля job в БД."""
    sets = ', '.join(f"{k}=%s" for k in kwargs)
    vals = list(kwargs.values()) + [job_id]
    cur = conn.cursor()
    cur.execute(f"UPDATE {SCHEMA}.import_jobs SET {sets}, updated_at=NOW() WHERE id=%s", vals)
    conn.commit()
    cur.close()


def _run_import_job(job_id: int):
    """Фоновый поток: скачивает XLSX и импортирует в market_listings батчами."""
    BATCH = 500
    MIN_PRICE_SALE, MAX_PRICE_SALE = 500_000, 5_000_000_000
    MIN_PRICE_RENT, MAX_PRICE_RENT = 5_000, 10_000_000
    MIN_AREA, MAX_AREA = 1, 200_000

    conn = psycopg2.connect(os.environ['DATABASE_URL'])
    try:
        # Загружаем параметры job
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute(f"SELECT * FROM {SCHEMA}.import_jobs WHERE id = %s", (job_id,))
        job = dict(cur.fetchone())
        cur.close()

        file_url = job['file_url']
        source = job['source']
        replace = job['replace_existing']

        _job_update(conn, job_id, status='downloading')

        # Скачиваем файл
        req = urllib.request.Request(file_url, headers={'User-Agent': 'Mozilla/5.0'})
        with urllib.request.urlopen(req, timeout=180) as r:
            raw_bytes = r.read()

        _job_update(conn, job_id, status='parsing')

        # Открываем XLSX потоково
        wb = openpyxl.load_workbook(io.BytesIO(raw_bytes), read_only=True, data_only=True)
        ws = wb.active

        # Читаем заголовок
        rows_iter = ws.iter_rows(values_only=True)
        header_row = next(rows_iter, None)
        if not header_row:
            _job_update(conn, job_id, status='error', error_msg='Файл пустой')
            wb.close()
            return

        header = [str(c).strip().lower() if c else '' for c in header_row]

        def find_col(*candidates) -> int:
            for cand in candidates:
                for i, h in enumerate(header):
                    if cand in h:
                        return i
            return -1

        col_price   = find_col('цена', 'price', 'стоимость')
        col_area    = find_col('площадь', 'area', 'кв.м', 'кв м', 'площ')
        col_deal    = find_col('тип сделки', 'тип объявления', 'deal', 'сделка', 'операция')
        col_cat     = find_col('категория', 'тип объекта', 'вид объекта', 'category', 'тип недвижимости', 'назначение')
        col_addr    = find_col('адрес', 'address', 'местоположение')
        col_dist    = find_col('район', 'district', 'округ')
        col_title   = find_col('название', 'заголовок', 'title', 'наименование')
        col_floor   = find_col('этаж', 'floor')
        col_tfloors = find_col('этажность', 'этажей', 'total_floor')
        col_url     = find_col('url', 'ссылка', 'link', 'объявление')
        col_ext_id  = find_col('id объявления', 'внешний id', 'id', 'номер объявления')
        col_desc    = find_col('описание', 'description', 'комментарий')
        col_ppm2    = find_col('цена за м', 'price_per_m', 'цена/м', 'руб/м')

        if col_price < 0 or col_area < 0:
            _job_update(conn, job_id, status='error',
                        error_msg=f'Не найдены колонки Цена/Площадь. Заголовок: {header[:20]}')
            wb.close()
            return

        # Если replace — удаляем старые записи
        if replace:
            cur2 = conn.cursor()
            cur2.execute(f"DELETE FROM {SCHEMA}.market_listings WHERE source = %s", (source,))
            conn.commit()
            cur2.close()

        _job_update(conn, job_id, status='running')

        rows_done = 0
        inserted = 0
        updated = 0
        skipped = 0
        cat_counts = {}
        seen_keys = set()
        batch = []

        def flush_batch(b):
            nonlocal inserted, updated
            cur3 = conn.cursor()
            for rec in b:
                ext_id = str(rec.get('external_id') or '')[:200] or \
                         f"xlsx_{source}_{rec.get('address','')[:40]}_{int(rec.get('area') or 0)}"
                cur3.execute(
                    f"INSERT INTO {SCHEMA}.market_listings "
                    f"(source, external_id, url, title, category, deal_type, price, price_per_m2, "
                    f"area, address, district, floor, total_floors, description, scraped_at) "
                    f"VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,NOW()) "
                    f"ON CONFLICT (source, external_id) DO UPDATE SET "
                    f"price=%s, price_per_m2=%s, area=%s, category=%s, deal_type=%s, "
                    f"address=%s, district=%s, title=%s, scraped_at=NOW()",
                    (
                        rec['source'], ext_id, rec.get('url'), rec.get('title'),
                        rec['category'], rec['deal_type'], rec.get('price'), rec.get('price_per_m2'),
                        rec.get('area'), rec.get('address'), rec.get('district'),
                        rec.get('floor'), rec.get('total_floors'), rec.get('description'),
                        rec.get('price'), rec.get('price_per_m2'), rec.get('area'),
                        rec['category'], rec['deal_type'],
                        rec.get('address'), rec.get('district'), rec.get('title'),
                    )
                )
                if cur3.rowcount == 1:
                    inserted += 1
                else:
                    updated += 1
            conn.commit()
            cur3.close()

        for row in rows_iter:
            rows_done += 1

            def cell(idx):
                if idx < 0 or idx >= len(row):
                    return None
                return row[idx]

            price = _parse_float_m(cell(col_price))
            area  = _parse_float_m(cell(col_area))
            deal  = _map_deal_m(str(cell(col_deal) or 'продажа'))
            title_v = str(cell(col_title) or '').strip() if col_title >= 0 else ''
            cat_raw = str(cell(col_cat) or '').strip()
            category = _map_obj_type_m(cat_raw, title_v)
            address  = str(cell(col_addr) or '').strip()
            district = str(cell(col_dist) or '').strip()[:200]
            floor_v  = int(_parse_float_m(cell(col_floor))) or None
            tfloors_v = int(_parse_float_m(cell(col_tfloors))) or None
            url_v   = str(cell(col_url) or '').strip() or None
            ext_id  = str(cell(col_ext_id) or '').strip() or None
            desc    = str(cell(col_desc) or '').strip()[:1000] or None
            ppm2_r  = _parse_float_m(cell(col_ppm2)) if col_ppm2 >= 0 else None
            ppm2    = ppm2_r if ppm2_r else _ppm2_m(price, area)

            # Фильтры
            if deal == 'sale' and not (MIN_PRICE_SALE <= price <= MAX_PRICE_SALE):
                skipped += 1
                continue
            if deal == 'rent' and not (MIN_PRICE_RENT <= price <= MAX_PRICE_RENT):
                skipped += 1
                continue
            if not (MIN_AREA <= area <= MAX_AREA):
                skipped += 1
                continue

            dk = f"{address or ext_id or f'r{rows_done}'}_{int(area)}"
            if dk in seen_keys:
                skipped += 1
                continue
            seen_keys.add(dk)

            cat_counts[category] = cat_counts.get(category, 0) + 1
            batch.append({
                'source': source,
                'external_id': ext_id,
                'url': url_v,
                'title': title_v[:500] or None,
                'category': category,
                'deal_type': deal,
                'price': int(price) if price else None,
                'price_per_m2': ppm2,
                'area': area or None,
                'address': address[:500] or None,
                'district': district or None,
                'floor': floor_v,
                'total_floors': tfloors_v,
                'description': desc,
            })

            if len(batch) >= BATCH:
                flush_batch(batch)
                batch = []
                # Обновляем прогресс каждые BATCH строк
                _job_update(conn, job_id,
                            rows_done=rows_done,
                            rows_inserted=inserted,
                            rows_updated=updated,
                            rows_skipped=skipped)

        # Последний батч
        if batch:
            flush_batch(batch)

        wb.close()

        _job_update(conn, job_id,
                    status='done',
                    rows_done=rows_done,
                    rows_inserted=inserted,
                    rows_updated=updated,
                    rows_skipped=skipped,
                    category_breakdown=json.dumps(cat_counts, ensure_ascii=False))

    except Exception as e:
        try:
            _job_update(conn, job_id, status='error', error_msg=str(e)[:500])
        except Exception:
            pass
    finally:
        conn.close()

COL_MAP = {
    'Отдельно стоящие здания за м2':        'standalone',
    'Производственные помещения за м2':      'industrial',
    'Торговые помещения и площади за м2':    'retail',
    'Помещения общепита за м2':              'catering',
    'Помещение свободного назначения за м2': 'free_purpose',
    'Складские помещения и комплексы за м2': 'warehouse',
    'Офисные помещения за м2':               'office',
}

RENT_FILE_KEYS = {
    '3f76641e-b047-4808-b575-4e245201e491.xlsx',
    'c1a2e6d7-4c98-4c21-9c9c-a6a8f264a839.xlsx',
    '5b5762ac-c687-4578-820f-4dcfbc8a6f23.xlsx',
}


def _parse_date(s: str):
    for fmt in ('%d.%m.%Y', '%Y-%m-%d', '%d/%m/%Y'):
        try:
            return datetime.strptime(s.strip(), fmt).date()
        except Exception:
            pass
    return None


def _is_rent(url: str) -> bool:
    return any(k in url for k in RENT_FILE_KEYS)


def handler(event: dict, context) -> dict:
    """Читает xlsx файлы с CDN и импортирует данные в price_history_biweekly или market_listings."""
    if event.get('httpMethod') == 'OPTIONS':
        return {'statusCode': 200, 'headers': {'Access-Control-Allow-Origin': '*', 'Access-Control-Allow-Methods': 'GET, POST, OPTIONS', 'Access-Control-Allow-Headers': 'Content-Type', 'Access-Control-Max-Age': '86400'}, 'body': ''}

    body = json.loads(event.get('body') or '{}')

    CORS_H = {'Access-Control-Allow-Origin': '*', 'Content-Type': 'application/json'}

    # ── Старт фонового импорта: создаём job, запускаем через threading ──────
    if body.get('action') == 'import_market_start':
        file_url = body.get('file_url', '').strip()
        source   = body.get('source', 'xlsx')[:50]
        replace  = bool(body.get('replace', False))
        if not file_url:
            return {'statusCode': 400, 'headers': CORS_H, 'body': json.dumps({'error': 'Укажите file_url'}, ensure_ascii=False)}
        conn = psycopg2.connect(os.environ['DATABASE_URL'])
        cur = conn.cursor()
        cur.execute(
            f"INSERT INTO {SCHEMA}.import_jobs (file_url, source, replace_existing, status) "
            f"VALUES (%s, %s, %s, 'pending') RETURNING id",
            (file_url, source, replace)
        )
        job_id = cur.fetchone()[0]
        conn.commit()
        cur.close()
        conn.close()
        # Запускаем фоновый поток
        import threading
        t = threading.Thread(target=_run_import_job, args=(job_id,), daemon=True)
        t.start()
        return {'statusCode': 200, 'headers': CORS_H, 'body': json.dumps({'job_id': job_id, 'status': 'pending'}, ensure_ascii=False)}

    # ── Статус job ────────────────────────────────────────────────────────────
    if body.get('action') == 'import_market_status':
        job_id = body.get('job_id')
        if not job_id:
            return {'statusCode': 400, 'headers': CORS_H, 'body': json.dumps({'error': 'job_id required'}, ensure_ascii=False)}
        conn = psycopg2.connect(os.environ['DATABASE_URL'])
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute(f"SELECT * FROM {SCHEMA}.import_jobs WHERE id = %s", (job_id,))
        row = cur.fetchone()
        cur.close()
        conn.close()
        if not row:
            return {'statusCode': 404, 'headers': CORS_H, 'body': json.dumps({'error': 'job not found'}, ensure_ascii=False)}
        return {'statusCode': 200, 'headers': CORS_H, 'body': json.dumps(dict(row), ensure_ascii=False, default=str)}

    # ── Список всех job-ов ────────────────────────────────────────────────────
    if body.get('action') == 'import_market_list':
        conn = psycopg2.connect(os.environ['DATABASE_URL'])
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute(f"SELECT * FROM {SCHEMA}.import_jobs ORDER BY created_at DESC LIMIT 20")
        rows = [dict(r) for r in cur.fetchall()]
        cur.close()
        conn.close()
        return {'statusCode': 200, 'headers': CORS_H, 'body': json.dumps(rows, ensure_ascii=False, default=str)}

    if body.get('action') == 'cleanup_biweekly':
        conn = psycopg2.connect(os.environ['DATABASE_URL'])
        cur = conn.cursor()
        keep_suffixes = ('_trend', '_peak', '_min',
                         '_2019', '_2020', '_2021', '_2022', '_2023', '_2024', '_2025', '_2026',
                         '_2019_2020_yoy', '_2020_2021_yoy', '_2021_2022_yoy',
                         '_2022_2023_yoy', '_2023_2024_yoy', '_2024_2025_yoy', '_2025_2026_yoy')
        cats = ('catering', 'free_purpose', 'industrial', 'office', 'retail', 'standalone', 'warehouse')
        dts = ('sale', 'rent')
        keep_keys = set()
        for cat in cats:
            for dt in dts:
                for sfx in keep_suffixes:
                    keep_keys.add(f'biweekly_{cat}_{dt}{sfx}')
        cur.execute(f"SELECT key FROM {SCHEMA}.ai_memory WHERE key LIKE 'biweekly_%'")
        all_keys = [r[0] for r in cur.fetchall()]
        to_delete = [k for k in all_keys if k not in keep_keys]
        deleted = 0
        for k in to_delete:
            cur.execute(f"DELETE FROM {SCHEMA}.ai_memory WHERE key = %s", (k,))
            deleted += cur.rowcount
        conn.commit()
        cur.close()
        conn.close()
        return {
            'statusCode': 200,
            'headers': {'Access-Control-Allow-Origin': '*', 'Content-Type': 'application/json'},
            'body': json.dumps({'deleted': deleted, 'kept': len(keep_keys), 'checked': len(all_keys)}, ensure_ascii=False),
        }

    urls = body.get('urls', [])
    preview_only = body.get('preview_only', False)

    all_rows = []
    errors = []

    for url in urls:
        deal_type = 'rent' if _is_rent(url) else 'sale'
        try:
            req = urllib.request.Request(url, headers={'User-Agent': 'Mozilla/5.0'})
            with urllib.request.urlopen(req, timeout=30) as r:
                data = r.read()
            wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                wb.close()
                continue
            header = [str(c).strip() if c else '' for c in rows[0]]

            data_cols = []
            for i, h in enumerate(header):
                if h in ('Даты', 'Изменение', ''):
                    continue
                category = COL_MAP.get(h, h.lower().replace(' ', '_').replace('/', '_'))
                data_cols.append((i, category))

            for row in rows[1:]:
                if not row or not row[0]:
                    continue
                date_val = _parse_date(str(row[0]))
                if not date_val:
                    continue
                for col_idx, category in data_cols:
                    if col_idx >= len(row) or row[col_idx] is None:
                        continue
                    try:
                        price = float(str(row[col_idx]).replace(',', '.').replace(' ', ''))
                    except Exception:
                        continue
                    change = None
                    next_idx = col_idx + 1
                    if next_idx < len(row) and row[next_idx]:
                        try:
                            change = float(str(row[next_idx]).replace('%', '').replace('+', '').strip())
                        except Exception:
                            pass
                    all_rows.append({
                        'date': date_val, 'category': category,
                        'deal_type': deal_type, 'price': price, 'change': change,
                    })
            wb.close()
        except Exception as e:
            errors.append({'url': url, 'error': str(e)})

    if preview_only:
        return {
            'statusCode': 200,
            'headers': {'Access-Control-Allow-Origin': '*', 'Content-Type': 'application/json'},
            'body': json.dumps({
                'total_rows': len(all_rows),
                'errors': errors,
                'sample': [{'date': str(r['date']), 'category': r['category'], 'deal_type': r['deal_type'], 'price': r['price']} for r in all_rows[:20]],
            }, ensure_ascii=False),
        }

    conn = psycopg2.connect(os.environ['DATABASE_URL'])
    cur = conn.cursor()
    inserted = 0
    skipped = 0
    for r in all_rows:
        try:
            cur.execute(
                f"INSERT INTO {SCHEMA}.price_history_biweekly "
                f"(date_recorded, category, deal_type, price_per_m2, change_pct, source) "
                f"VALUES (%s, %s, %s, %s, %s, 'xlsx_import') "
                f"ON CONFLICT (date_recorded, category, deal_type) DO NOTHING",
                (r['date'], r['category'], r['deal_type'], r['price'], r['change'])
            )
            inserted += cur.rowcount
        except Exception as e:
            skipped += 1
            errors.append({'row': str(r['date']) + '/' + r['category'], 'error': str(e)[:100]})
    conn.commit()
    cur.close()
    conn.close()

    return {
        'statusCode': 200,
        'headers': {'Access-Control-Allow-Origin': '*', 'Content-Type': 'application/json'},
        'body': json.dumps({'success': True, 'inserted': inserted, 'skipped': skipped, 'total_parsed': len(all_rows), 'errors': errors[:10]}, ensure_ascii=False),
    }